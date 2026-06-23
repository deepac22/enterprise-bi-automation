import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.drawing.image import Image
from datetime import datetime
import os
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import io

us_state_abbrev_to_full = {
    'AL': 'Alabama', 'AK': 'Alaska', 'AZ': 'Arizona', 'AR': 'Arkansas',
    'CA': 'California', 'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware',
    'FL': 'Florida', 'GA': 'Georgia', 'HI': 'Hawaii', 'ID': 'Idaho',
    'IL': 'Illinois', 'IN': 'Indiana', 'IA': 'Iowa', 'KS': 'Kansas',
    'KY': 'Kentucky', 'LA': 'Louisiana', 'ME': 'Maine', 'MD': 'Maryland',
    'MA': 'Massachusetts', 'MI': 'Michigan', 'MN': 'Minnesota', 'MS': 'Mississippi',
    'MO': 'Missouri', 'MT': 'Montana', 'NE': 'Nebraska', 'NV': 'Nevada',
    'NH': 'New Hampshire', 'NJ': 'New Jersey', 'NM': 'New Mexico', 'NY': 'New York',
    'NC': 'North Carolina', 'ND': 'North Dakota', 'OH': 'Ohio', 'OK': 'Oklahoma',
    'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island', 'SC': 'South Carolina',
    'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah',
    'VT': 'Vermont', 'VA': 'Virginia', 'WA': 'Washington', 'WV': 'West Virginia',
    'WI': 'Wisconsin', 'WY': 'Wyoming'
}

def load_data(file_path):
    df = pd.read_csv(file_path, nrows=100000, low_memory=False)
    valid_statuses = ['Fully Paid', 'Charged Off', 'Current', 'Default', 'Late (31-120 days)']
    df = df[df['loan_status'].isin(valid_statuses)]
    df = df.dropna(subset=['addr_state', 'loan_amnt', 'int_rate'])
    return df

def calculate_state_metrics(df):
    state_agg = df.groupby('addr_state').agg(
        total_loans=('loan_amnt', 'count'),
        avg_amount=('loan_amnt', 'mean'),
        delinquency_rate=('loan_status', lambda x: (x.isin(['Charged Off', 'Default', 'Late (31-120 days)']).sum() / len(x)) * 100)
    ).reset_index().round(2)
    state_agg = state_agg.sort_values('delinquency_rate', ascending=False)
    return state_agg

def calculate_kpis(df):
    total_loans = len(df)
    total_value = df['loan_amnt'].sum()
    avg_rate = df['int_rate'].mean()
    delinquent_statuses = ['Charged Off', 'Default']
    del_rate = (df['loan_status'].isin(delinquent_statuses).sum() / total_loans) * 100
    at_risk = df[df['loan_status'].isin(delinquent_statuses)].shape[0]
    return {
        'total_loans': total_loans,
        'total_value': total_value,
        'avg_rate': avg_rate,
        'del_rate': del_rate,
        'at_risk': at_risk
    }

def apply_heatmap(worksheet, data_range):
    color_scale = ColorScaleRule(
        start_type='min', start_color='00FF00',
        mid_type='percentile', mid_value=50, mid_color='FFEB3B',
        end_type='max', end_color='FF0000'
    )
    worksheet.conditional_formatting.add(data_range, color_scale)

def generate_cover_page(workbook):
    ws = workbook.active
    ws.title = "Cover"
    ws.merge_cells('A1:O1')
    ws.merge_cells('A2:O2')
    ws.merge_cells('A3:O3')
    
    header_fill = PatternFill(start_color='0A1E3D', end_color='0A1E3D', fill_type='solid')
    for row in range(1, 4):
        for col in range(1, 16):
            ws.cell(row=row, column=col).fill = header_fill
    
    ws.merge_cells('A4:O4')
    ws['A4'] = ''
    for col in range(1, 16):
        ws.cell(row=4, column=col).fill = PatternFill(start_color='D4AF37', end_color='D4AF37', fill_type='solid')
        ws.row_dimensions[4].height = 4
    
    ws.merge_cells('A6:O6')
    ws['A6'] = 'GLOBAL LENDING PORTFOLIO REPORT'
    ws['A6'].font = Font(size=28, bold=True, color='0A1E3D', name='Calibri')
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[6].height = 45
    
    ws.merge_cells('A7:O7')
    ws['A7'] = 'Enterprise Risk & Performance Analysis'
    ws['A7'].font = Font(size=14, color='7F8C8D', name='Calibri', italic=True)
    ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[7].height = 25
    
    ws.merge_cells('A9:O9')
    ws['A9'] = '----------------------------------------------------------------------------------'
    ws['A9'].font = Font(color='D4AF37', size=12)
    ws['A9'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A11:O11')
    ws['A11'] = 'Prepared For:  Executive Leadership'
    ws['A11'].font = Font(size=14, color='34495E', bold=True)
    ws['A11'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A12:O12')
    ws['A12'] = 'Prepared By:  Analytics & Risk Intelligence Team'
    ws['A12'].font = Font(size=12, color='7F8C8D')
    ws['A12'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A14:O14')
    ws['A14'] = f'DATE: {datetime.now().strftime("%B %d, %Y").upper()}'
    ws['A14'].font = Font(size=16, bold=True, color='0A1E3D')
    ws['A14'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[14].height = 35
    
    ws.merge_cells('A16:O16')
    ws['A16'] = 'CONFIDENTIAL - FOR INTERNAL USE ONLY'
    ws['A16'].font = Font(size=14, bold=True, color='E74C3C')
    ws['A16'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A20:O20')
    ws['A20'] = ''
    for col in range(1, 16):
        ws.cell(row=20, column=col).fill = PatternFill(start_color='D4AF37', end_color='D4AF37', fill_type='solid')
        ws.row_dimensions[20].height = 3
    
    for col in range(1, 16):
        ws.column_dimensions[chr(64 + col)].width = 12
    
    return ws

def generate_kpi_dashboard(workbook, kpis):
    ws = workbook.create_sheet("KPI Dashboard")
    
    ws.merge_cells('A1:D1')
    ws['A1'] = 'KPI DASHBOARD'
    ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0A1E3D', end_color='0A1E3D', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    kpi_data = [
        ['Total Loans', f"{kpis['total_loans']:,.0f}", 'Loans', 'FF1a5276'],
        ['Portfolio Value', f"${kpis['total_value']:,.0f}", 'USD', 'FF1a5276'],
        ['Delinquency Rate', f"{kpis['del_rate']:.2f}%", 'Risk', 'FFc0392b' if kpis['del_rate'] > 10 else 'FF27ae60'],
        ['At-Risk Loans', f"{kpis['at_risk']:,.0f}", 'Accounts', 'FFc0392b' if kpis['at_risk'] > 500 else 'FF27ae60']
    ]
    
    row_start = 3
    for i, (label, value, sub, color) in enumerate(kpi_data):
        r = row_start + (i // 2) * 4
        c = 1 + (i % 2) * 2
        
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
        cell_label = ws.cell(row=r, column=c)
        cell_label.value = label
        cell_label.font = Font(size=12, bold=True, color='7F8C8D')
        cell_label.alignment = Alignment(horizontal='center')
        
        ws.merge_cells(start_row=r+1, start_column=c, end_row=r+1, end_column=c+1)
        cell_value = ws.cell(row=r+1, column=c)
        cell_value.value = value
        cell_value.font = Font(size=20, bold=True, color=color)
        cell_value.alignment = Alignment(horizontal='center')
        
        ws.merge_cells(start_row=r+2, start_column=c, end_row=r+2, end_column=c+1)
        cell_sub = ws.cell(row=r+2, column=c)
        cell_sub.value = sub
        cell_sub.font = Font(size=10, color='95A5A6')
        cell_sub.alignment = Alignment(horizontal='center')
        
        for row in range(r, r+3):
            for col in range(c, c+2):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin', color='BDC3C7'),
                    right=Side(style='thin', color='BDC3C7'),
                    top=Side(style='thin', color='BDC3C7'),
                    bottom=Side(style='thin', color='BDC3C7')
                )
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    for row in range(3, 12):
        ws.row_dimensions[row].height = 25
    
    return ws

def generate_heatmap(workbook, state_agg):
    ws = workbook.create_sheet("US Risk Heatmap")
    ws.append(['State', 'Delinquency Rate (%)', 'Total Loans'])
    for _, row in state_agg.iterrows():
        ws.append([f"{row['state_full']} ({row['addr_state']})", row['delinquency_rate'], row['total_loans']])
    data_range = f"B2:B{len(state_agg) + 1}"
    apply_heatmap(ws, data_range)
    for col in range(1, 4):
        ws.column_dimensions[chr(64 + col)].width = 30
    return ws

def generate_raw_data(workbook, df):
    ws = workbook.create_sheet("Raw Data")
    headers = ['ID', 'State', 'Amount', 'Rate', 'Status', 'Grade']
    ws.append(headers)
    for i, row in df.head(1000).iterrows():
        state_full = us_state_abbrev_to_full.get(row['addr_state'], row['addr_state'])
        ws.append([i, f"{state_full} ({row['addr_state']})", row['loan_amnt'], row['int_rate'], row['loan_status'], row.get('grade', 'N/A')])
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 20
    return ws

def generate_executive_summary(workbook, state_agg, kpis):
    ws = workbook.create_sheet("Executive Summary")
    
    worst = state_agg.iloc[0]
    best = state_agg.iloc[-1]
    
    ws['A1'] = 'EXECUTIVE SUMMARY'
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0A1E3D', end_color='0A1E3D', fill_type='solid')
    ws.merge_cells('A1:D1')
    
    ws['A3'] = 'Portfolio Overview'
    ws['A3'].font = Font(size=14, bold=True)
    ws['A4'] = f'Total Loans: {kpis["total_loans"]:,.0f}'
    ws['A5'] = f'Total Portfolio Value: ${kpis["total_value"]:,.0f}'
    ws['A6'] = f'Overall Delinquency Rate: {kpis["del_rate"]:.2f}%'
    ws['A7'] = f'Average Interest Rate: {kpis["avg_rate"]:.2f}%'
    
    ws['A9'] = 'Risk Assessment'
    ws['A9'].font = Font(size=14, bold=True)
    ws['A10'] = f'Highest Risk State: {worst["state_full"]} ({worst["addr_state"]}) - {worst["delinquency_rate"]:.2f}%'
    ws['A11'] = f'Lowest Risk State: {best["state_full"]} ({best["addr_state"]}) - {best["delinquency_rate"]:.2f}%'
    ws['A12'] = f'At-Risk Loans: {kpis["at_risk"]:,.0f}'
    
    ws['A14'] = 'Strategic Recommendations'
    ws['A14'].font = Font(size=14, bold=True)
    ws['A15'] = '1. Review underwriting criteria for states with delinquency rates exceeding 10%.'
    ws['A16'] = '2. Consider portfolio rebalancing toward lower-risk states and products.'
    ws['A17'] = '3. Implement early warning systems for at-risk loan segments.'
    
    for row in range(3, 18):
        ws.row_dimensions[row].height = 22
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    return ws


def generate_state_chart(workbook, state_agg):
    ws = workbook.create_sheet("State Chart")
    
    # Write data for reference
    ws.append(['State', 'Delinquency Rate'])
    for _, row in state_agg.head(15).iterrows():
        ws.append([f"{row['state_full']} ({row['addr_state']})", row['delinquency_rate']])
    
    # --- Generate the chart using matplotlib ---
    # Extract data
    states = [f"{row['state_full']} ({row['addr_state']})" for _, row in state_agg.head(15).iterrows()]
    rates = [row['delinquency_rate'] for _, row in state_agg.head(15).iterrows()]
    
    # Create figure
    fig, ax = plt.subplots(figsize=(12, 6))
    
    # Dark blue bars with gold borders
    bars = ax.bar(states, rates, color='#0A1E3D', edgecolor='#D4AF37', linewidth=2)
    
    # Add value labels on top of bars
    for bar in bars:
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height + 0.3,
                f'{height:.1f}%', ha='center', va='bottom', fontsize=9, fontweight='bold', color='#0A1E3D')
    
    # Titles and labels
    ax.set_title('Top 15 States by Delinquency Rate', fontsize=16, fontweight='bold', color='#0A1E3D', pad=20)
    ax.set_xlabel('State', fontsize=12, fontweight='bold', color='#34495E')
    ax.set_ylabel('Delinquency Rate (%)', fontsize=12, fontweight='bold', color='#34495E')
    
    # Rotate x-axis labels for readability
    ax.set_xticklabels(states, rotation=45, ha='right', fontsize=9, color='#34495E')
    ax.tick_params(axis='y', labelsize=9, colors='#34495E')
    
    # Remove top and right spines
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#BDC3C7')
    ax.spines['bottom'].set_color('#BDC3C7')
    
    # Grid lines (horizontal)
    ax.yaxis.grid(True, linestyle='--', alpha=0.6, color='#BDC3C7')
    ax.set_axisbelow(True)
    
    # Tight layout
    plt.tight_layout()
    
    # Save the chart to a BytesIO object
    img_data = io.BytesIO()
    plt.savefig(img_data, format='png', dpi=150, bbox_inches='tight')
    img_data.seek(0)
    plt.close()
    
    # --- Insert the image into Excel ---
    img = Image(img_data)
    # Scale the image to fit nicely
    img.width = 800
    img.height = 400
    ws.add_image(img, 'D2')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['D'].width = 50
    
    return ws

def main():
    print("Loading real loan data...")
    df = load_data('real_loan_data.csv')
    print("Calculating metrics...")
    state_agg = calculate_state_metrics(df)
    state_agg['state_full'] = state_agg['addr_state'].map(us_state_abbrev_to_full)
    kpis = calculate_kpis(df)
    print("Generating Excel report...")
    wb = Workbook()
    generate_cover_page(wb)
    generate_kpi_dashboard(wb, kpis)
    generate_heatmap(wb, state_agg)
    generate_raw_data(wb, df)
    generate_executive_summary(wb, state_agg, kpis)
    generate_state_chart(wb, state_agg)
    os.makedirs('reports', exist_ok=True)
    filename = f"reports/Global_Lending_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"Report saved: {filename}")
    try:
        from upload_to_google_sheets import upload_excel_to_google_sheets
        SPREADSHEET_ID = "1Sm1-Bl64G26hnFHHMfL14I_NTd6JXDuUBs_HZFmF2G0"
        upload_excel_to_google_sheets(filename, SPREADSHEET_ID, 'Lending Data')
        print("Uploaded to Google Sheets.")
    except Exception as e:
        print(f"Google Sheets upload failed: {e}")

if __name__ == "__main__":
    main()